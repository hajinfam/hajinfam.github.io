# pipeline_allin_append.py (FINAL + VERIFY)
# 목표:
# 1) excel_re_3type.py 실행(번호 선택 포함)
# 2) re_xlsx/products_MMDD.xlsx(이번 실행에서 새로 생성된 파일) 선택
# 3) 그 파일의 2행~끝을 읽어 products.xlsx 맨 아래에 append
# 4) append된 행들의 no를 기존 마지막 no+1부터 순차 부여
# 5) products.xlsx의 productUrl/productLink는 "반드시 link 컬럼 값"으로만 채움 (어필리에이트 규칙)
# 6) append 직후 마지막 N행을 검사하여 link 강제 반영이 맞는지 확인
# 7) allin.py 실행

import os
import sys
import time
import glob
import subprocess
from datetime import datetime
from typing import List, Dict, Tuple, Optional

from openpyxl import load_workbook


# -----------------------------
# Paths
# -----------------------------
REPO_ROOT = r"H:\hajinfam.github.io"
RE_XLSX_DIR = os.path.join(REPO_ROOT, "re_xlsx")

EXCEL_RE_3TYPE = os.path.join(REPO_ROOT, "excel_re_3type.py")
ALLIN_PY = os.path.join(REPO_ROOT, "allin.py")

TARGET_PRODUCTS_XLSX = os.path.join(REPO_ROOT, "products.xlsx")

# 검증 시 마지막 몇 행을 확인할지
VERIFY_TAIL_N = 10


# -----------------------------
# Utilities
# -----------------------------
def list_products_mmdd_files(folder: str) -> List[str]:
    files = sorted(glob.glob(os.path.join(folder, "products_*.xlsx")))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]  # temp 제외
    return files


def run_py(path: str, cwd: str):
    if not os.path.exists(path):
        raise FileNotFoundError(f"[ERROR] file not found: {path}")
    subprocess.check_call([sys.executable, path], cwd=cwd)


def pick_new_file(before: set, after: List[str]) -> str:
    new_files = [f for f in after if f not in before]
    if len(new_files) == 1:
        return new_files[0]
    if len(new_files) == 0:
        raise RuntimeError(
            "[ERROR] 새로 생성된 products_*.xlsx를 찾지 못했습니다.\n"
            "- excel_re_3type.py가 re_xlsx 폴더에 products_MMDD.xlsx를 '새 파일'로 생성하는지 확인하세요.\n"
            "- 저장 위치/파일명 규칙이 다르면 이 에러가 납니다."
        )
    raise RuntimeError(f"[ERROR] 새로 생성된 products_*.xlsx가 2개 이상입니다(모호): {new_files}")


def load_first_sheet(path: str) -> Tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(path)
    ws = wb.active

    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append("" if v is None else str(v).strip())

    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        d: Dict[str, str] = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(row=r, column=c).value
            s = "" if v is None else str(v).strip()
            d[h] = s
            if s:
                empty = False
        if not empty:
            rows.append(d)

    return headers, rows


def find_header_index(headers: List[str], name: str) -> Optional[int]:
    name_low = name.strip().lower()
    for i, h in enumerate(headers):
        if (h or "").strip().lower() == name_low:
            return i
    return None


def get_last_no_from_target(ws, target_headers: List[str]) -> int:
    no_col_idx = find_header_index(target_headers, "no")
    if no_col_idx is None:
        raise RuntimeError(
            "[ERROR] products.xlsx 헤더에 'no' 컬럼이 없습니다.\n"
            "인스타/인스타깃 동기화를 위해 no가 필수입니다."
        )
    no_col = no_col_idx + 1  # 1-based

    last = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=no_col).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        try:
            n = int(float(s))
            if n > last:
                last = n
        except:
            continue
    return last


def append_rows_products_only_link(
    target_products_path: str,
    src_headers: List[str],
    src_rows: List[Dict[str, str]],
) -> Tuple[int, int]:
    """
    강제 규칙:
      - src의 productUrl/productDescriptionUrl은 절대 읽지 않음
      - src의 link만 사용
      - target의 productUrl/productLink는 반드시 link 값으로 채움
    """
    if not os.path.exists(target_products_path):
        raise FileNotFoundError(f"[ERROR] products.xlsx not found: {target_products_path}")

    if "link" not in src_headers:
        raise RuntimeError(
            "[ERROR] products_MMDD.xlsx에 'link' 컬럼이 없습니다.\n"
            "규칙상 반드시 link만 사용해야 합니다."
        )

    twb = load_workbook(target_products_path)
    tws = twb.active

    target_headers: List[str] = []
    for c in range(1, tws.max_column + 1):
        v = tws.cell(row=1, column=c).value
        target_headers.append("" if v is None else str(v).strip())

    has_productUrl = find_header_index(target_headers, "productUrl") is not None
    has_productLink = find_header_index(target_headers, "productLink") is not None
    if not (has_productUrl or has_productLink):
        raise RuntimeError(
            "[ERROR] products.xlsx 헤더에 productUrl 또는 productLink 컬럼이 없습니다."
        )

    last_no = get_last_no_from_target(tws, target_headers)
    t_col = {h: i + 1 for i, h in enumerate(target_headers) if h}

    appended_rows = 0
    start_row = tws.max_row + 1

    for d in src_rows:
        # ★ 절대 다른 URL 컬럼을 읽지 않음
        link = (d.get("link") or "").strip()

        pname = (d.get("productName") or "").strip()
        main = (d.get("mainImage") or "").strip()
        sub = (d.get("subImage") or "").strip()

        if not (pname and main and link):
            continue
        if not (link.startswith("http://") or link.startswith("https://")):
            continue

        last_no += 1
        rr = start_row + appended_rows

        if "no" in t_col:
            tws.cell(row=rr, column=t_col["no"]).value = str(last_no)
        if "productName" in t_col:
            tws.cell(row=rr, column=t_col["productName"]).value = pname
        if "mainImage" in t_col:
            tws.cell(row=rr, column=t_col["mainImage"]).value = main
        if "subImage" in t_col:
            tws.cell(row=rr, column=t_col["subImage"]).value = sub

        # ★ 핵심: products.xlsx의 URL 컬럼은 무조건 link
        if "productUrl" in t_col:
            tws.cell(row=rr, column=t_col["productUrl"]).value = link
        if "productLink" in t_col:
            tws.cell(row=rr, column=t_col["productLink"]).value = link

        appended_rows += 1

    if appended_rows == 0:
        raise RuntimeError(
            "[ERROR] append할 유효 행이 없습니다.\n"
            "- productName/mainImage/link 누락 또는 link 형식 오류로 스킵되었을 수 있습니다."
        )

    twb.save(target_products_path)
    return appended_rows, last_no


def verify_tail_rows_products_link(
    target_products_path: str,
    appended_rows: int,
    expected_links_in_order: List[str],
    tail_n: int = 10,
):
    """
    append 직후 products.xlsx의 마지막 N행을 출력/검증:
    - productUrl 또는 productLink 값이 expected_links_in_order(append된 순서)와 정확히 일치해야 함
    - 하나라도 불일치하면 RuntimeError
    """
    wb = load_workbook(target_products_path)
    ws = wb.active

    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append("" if v is None else str(v).strip())
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    url_col_name = None
    if "productUrl" in col:
        url_col_name = "productUrl"
    elif "productLink" in col:
        url_col_name = "productLink"
    else:
        raise RuntimeError("[VERIFY ERROR] products.xlsx에 productUrl/productLink가 없습니다.")

    # 방금 append된 마지막 appended_rows 구간을 우선 검증
    last_row = ws.max_row
    first_appended_row = last_row - appended_rows + 1

    # 검증 대상: appended 구간 전체(정확성), 출력은 tail_n만
    failures = 0

    print("\n[VERIFY] 최근 append 구간 URL 강제 반영 검사:")
    print(f"- URL column used: {url_col_name}")
    print(f"- appended_rows: {appended_rows}")
    print(f"- appended range: {first_appended_row} ~ {last_row}")

    for i in range(appended_rows):
        rr = first_appended_row + i
        actual = ws.cell(row=rr, column=col[url_col_name]).value
        actual = "" if actual is None else str(actual).strip()
        expected = expected_links_in_order[i].strip()

        if actual != expected:
            failures += 1
            print(f"[MISMATCH] row={rr} expected(link) != actual({url_col_name})")
            print(f"  expected: {expected}")
            print(f"  actual  : {actual}")

    # tail 출력
    print(f"\n[TAIL] products.xlsx 마지막 {tail_n}행(요약):")
    start = max(2, ws.max_row - tail_n + 1)
    for rr in range(start, ws.max_row + 1):
        no_v = ws.cell(row=rr, column=col["no"]).value if "no" in col else ""
        name_v = ws.cell(row=rr, column=col.get("productName", 0)).value if "productName" in col else ""
        url_v = ws.cell(row=rr, column=col[url_col_name]).value
        no_s = "" if no_v is None else str(no_v).strip()
        name_s = "" if name_v is None else str(name_v).strip()
        url_s = "" if url_v is None else str(url_v).strip()
        print(f"- row {rr}: no={no_s} | productName={name_s[:30]} | {url_col_name}={url_s}")

    if failures:
        raise RuntimeError(f"[VERIFY FAILED] URL 강제 반영 불일치 {failures}건 발견. allin.py 실행을 중단합니다.")
    print("[VERIFY OK] URL 강제 반영이 정상입니다.")


def main():
    print("==================================================")
    print("ALLIN PIPELINE FINAL + VERIFY")
    print("- link 컬럼만 사용(어필리에이트 규칙 강제)")
    print("- products.xlsx의 productUrl/productLink는 반드시 link 값으로 채움")
    print("- no 자동 증가")
    print("Repo:", REPO_ROOT)
    print("Time:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("==================================================")

    if not os.path.exists(RE_XLSX_DIR):
        raise FileNotFoundError(f"[ERROR] re_xlsx dir not found: {RE_XLSX_DIR}")

    # STEP 1: 실행 전 products_MMDD 목록 스냅샷
    before = set(list_products_mmdd_files(RE_XLSX_DIR))

    # STEP 2: excel_re_3type 실행
    print("\n[STEP 1] excel_re_3type.py 실행 (번호 선택 포함)")
    run_py(EXCEL_RE_3TYPE, cwd=REPO_ROOT)
    time.sleep(0.3)

    # STEP 3: 이번 실행에서 새로 생성된 products_MMDD.xlsx 선택
    after = list_products_mmdd_files(RE_XLSX_DIR)
    picked = pick_new_file(before, after)
    print(f"\n[STEP 2] 이번 실행 결과 products 파일: {picked}")

    # STEP 4: 소스 읽기
    src_headers, src_rows = load_first_sheet(picked)
    if not src_rows:
        raise RuntimeError("[ERROR] 선택된 products_MMDD.xlsx에 데이터 행이 없습니다(2행~끝).")
    print(f"[INFO] source rows: {len(src_rows)}")

    # STEP 5: append를 위해 "유효행의 link 순서"를 미리 수집(검증용)
    # (append 함수와 동일한 기준으로 필터링)
    expected_links_in_order: List[str] = []
    for d in src_rows:
        link = (d.get("link") or "").strip()
        pname = (d.get("productName") or "").strip()
        main = (d.get("mainImage") or "").strip()
        if not (pname and main and link):
            continue
        if not (link.startswith("http://") or link.startswith("https://")):
            continue
        expected_links_in_order.append(link)

    # STEP 6: products.xlsx append
    print("\n[STEP 3] products.xlsx에 append (no 자동 증가 + link 강제 반영)")
    appended, new_last_no = append_rows_products_only_link(
        target_products_path=TARGET_PRODUCTS_XLSX,
        src_headers=src_headers,
        src_rows=src_rows,
    )
    print(f"[OK] appended: {appended}, new last no: {new_last_no}")

    # STEP 7: 검증(append된 구간 전체 불일치 검사 + tail 출력)
    print("\n[STEP 3.5] link 강제 반영 검증")
    verify_tail_rows_products_link(
        target_products_path=TARGET_PRODUCTS_XLSX,
        appended_rows=appended,
        expected_links_in_order=expected_links_in_order,
        tail_n=VERIFY_TAIL_N,
    )

    # STEP 8: allin 실행
    print("\n[STEP 4] allin.py 실행")
    run_py(ALLIN_PY, cwd=REPO_ROOT)

    print("\n[DONE] ALLIN pipeline completed successfully.")
    print("==================================================")


if __name__ == "__main__":
    main()
