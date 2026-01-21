import os
import sys
import time
import json
import traceback
from typing import Dict, Any, List, Optional

import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from openai import OpenAI
import gspread
from google.oauth2.service_account import Credentials


# =========================
# Google Sheet
# =========================
GSERVICE_KEY_PATH = "hajinfam-products-9996c020862f.json"
GOOGLE_SHEET_ID = "1oKLktbTJgCJaRVfZz76rPuUlt0QrPQvE4c3un4OTRJA"  # /d/{ID}/ 부분
GOOGLE_WORKSHEET_NAME = "insta"  # 탭 이름
GS_WRITE_MODE = "fill_empty"  # "fill_empty" 또는 "overwrite"

# =========================
# 설정
# =========================
EXCEL_PATH = "products.xlsx"
SHEET_NAME = None  # 예: "Sheet1" / None이면 첫 시트
OUTPUT_DIR = "site_products"
BASE_PATH = "/site_products"

REQUIRED_COLUMNS = ["no", "productName", "productUrl", "mainImage", "shortTitle", "productDescription"]
SEO_COLUMNS = ["seoTitle", "seoDescription", "keywords", "category"]

# 생성/재개 옵션
SKIP_EXISTING_MD = True          # 이미 md가 있으면 스킵 (재개 실행용)
STOP_AFTER: Optional[int] = None # 디버깅용: 특정 개수 생성 후 중단 (None이면 끝까지)

# OpenAI 옵션
OPENAI_MODEL = "gpt-4.1-mini"
OPENAI_TIMEOUT_SEC = 30
OPENAI_MAX_RETRIES = 3


# =========================
# OpenAI
# =========================
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise RuntimeError("OPENAI_API_KEY 가 설정되어 있지 않습니다. .env 를 확인하세요.")

# timeout을 클라이언트에 직접 부여 (네트워크 지연 시 '멈춤' 방지)
client = OpenAI(api_key=OPENAI_API_KEY, timeout=OPENAI_TIMEOUT_SEC)


def to_str(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def shorten(s: str, n: int) -> str:
    s = (s or "").strip()
    return s if len(s) <= n else s[:n].rstrip() + "…"


def html_escape(s: str) -> str:
    return (
        str(s).replace("&", "&amp;")
              .replace("<", "&lt;")
              .replace(">", "&gt;")
              .replace('"', "&quot;")
    )


def generate_seo_pack(name: str, desc: str, keywords: str = "", category: str = "") -> Dict[str, Any]:
    """
    OpenAI 호출:
    - timeout 적용됨
    - 재시도 적용됨
    - 실패 시 폴백 반환 (멈춤 방지)
    """
    system_msg = (
        "너는 한국 온라인 쇼핑/특가 링크 페이지용 SEO 카피라이터다.\n"
        "네이버 친화 규칙:\n"
        "- 과장/허위/선정적 표현 금지\n"
        "- seoTitle: 40~55자 권장, 핵심키워드+상품명+의도(추천/할인/구매) 1개 포함\n"
        "- seoDescription: 80~120자, 장점/용도/대상 중심 1~2문장\n"
        "- keywords: 콤마로 구분된 5~10개(너무 길지 않게)\n"
        "- category: 1~2단어 (예: 크리스마스/인테리어/가전)\n"
        "- intro 1~2문장(120자 이내)\n"
        "- bullets 3~5개\n"
        "- rows 표 스펙 5~8개(label/detail)\n"
        "- faq 2~3개(q/a)\n"
        "출력은 JSON만.\n"
        "{"
        "\"seoTitle\":\"...\","
        "\"seoDescription\":\"...\","
        "\"keywords\":\"키워드1,키워드2,...\","
        "\"category\":\"...\","
        "\"intro\":\"...\","
        "\"bullets\":[\"...\"],"
        "\"rows\":[{\"label\":\"...\",\"detail\":\"...\"}],"
        "\"faq\":[{\"q\":\"...\",\"a\":\"...\"}]"
        "}"
    )

    user_msg = (
        f"상품명: {name}\n"
        f"설명: {desc}\n"
        f"기존 키워드(있으면 참고): {keywords}\n"
        f"기존 카테고리(있으면 참고): {category}\n"
        "JSON만 출력."
    )

    last_err: Optional[Exception] = None
    text = ""

    for attempt in range(1, OPENAI_MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": system_msg},
                    {"role": "user", "content": user_msg},
                ],
                max_tokens=700,
            )
            text = (resp.choices[0].message.content or "").strip()
            break
        except Exception as e:
            last_err = e
            print(f"[WARN] OpenAI failed (attempt {attempt}/{OPENAI_MAX_RETRIES}): {e}", flush=True)
            time.sleep(2 * attempt)

    if not text:
        # 폴백 (OpenAI가 계속 실패해도 전체 파이프라인은 진행)
        return {
            "seoTitle": name,
            "seoDescription": shorten(desc, 110),
            "keywords": keywords,
            "category": category,
            "intro": shorten(desc, 120),
            "bullets": [],
            "rows": [],
            "faq": [],
        }

    # ```json ... ``` 방어
    if text.startswith("```"):
        text = text.strip("`").strip()
        lines = text.splitlines()
        if lines and "{" not in lines[0]:
            text = "\n".join(lines[1:]).strip()

    try:
        data = json.loads(text)
    except Exception:
        return {
            "seoTitle": name,
            "seoDescription": shorten(desc, 110),
            "keywords": keywords,
            "category": category,
            "intro": shorten(desc, 120),
            "bullets": [],
            "rows": [],
            "faq": [],
        }

    data["seoTitle"] = to_str(data.get("seoTitle")) or name
    data["seoDescription"] = to_str(data.get("seoDescription")) or shorten(desc, 110)
    data["keywords"] = to_str(data.get("keywords")) or keywords
    data["category"] = to_str(data.get("category")) or category
    data["intro"] = to_str(data.get("intro"))
    data["bullets"] = data.get("bullets") or []
    data["rows"] = data.get("rows") or []
    data["faq"] = data.get("faq") or []
    return data


def build_markdown(row: Dict[str, Any], pack: Dict[str, Any]) -> str:
    no = int(row["no"])
    name = to_str(row.get("productName"))
    short_title = to_str(row.get("shortTitle"))
    product_url = to_str(row.get("productUrl"))
    image_url = to_str(row.get("mainImage"))
    base_desc = to_str(row.get("productDescription"))

    seo_title = to_str(row.get("seoTitle")) or to_str(pack.get("seoTitle")) or (short_title or name)
    seo_desc = to_str(row.get("seoDescription")) or to_str(pack.get("seoDescription")) or shorten(base_desc, 110)
    keywords = to_str(row.get("keywords")) or to_str(pack.get("keywords"))
    category = to_str(row.get("category")) or to_str(pack.get("category"))

    permalink = f"{BASE_PATH}/product_{no:03d}/"

    intro = to_str(pack.get("intro"))
    bullets = pack.get("bullets") or []
    rows = pack.get("rows") or []
    faq = pack.get("faq") or []

    product_ld = f'''<script type="application/ld+json">
{{
  "@context": "https://schema.org",
  "@type": "Product",
  "name": "{html_escape(short_title or name)}",
  "image": ["{image_url}"],
  "description": "{html_escape(seo_desc)}",
  "brand": {{"@type":"Brand","name":"hajinfam"}},
  "url": "{product_url}"
}}
</script>'''

    lines: List[str] = []
    lines.append("---")
    lines.append(f'title: "{html_escape(seo_title)}"')
    lines.append(f'description: "{html_escape(seo_desc)}"')
    lines.append(f"product_no: {no}")
    lines.append(f'product_name: "{html_escape(name)}"')
    if category:
        lines.append(f'category: "{html_escape(category)}"')
    if keywords:
        lines.append(f'keywords: "{html_escape(keywords)}"')
    lines.append(f'product_url: "{product_url}"')
    lines.append(f'image_url: "{image_url}"')
    lines.append(f'permalink: "{permalink}"')
    lines.append("sitemap: show")
    lines.append("noindex: false")
    lines.append("---\n")

    lines.append(f"# {no}번. {html_escape(short_title or name)}\n")
    if image_url:
        lines.append(f"![{html_escape(short_title or name)}]({image_url})\n")

    lines.append("## 핵심 요약")
    lines.append(intro or "상품 핵심 정보를 간단히 정리했습니다. 자세한 조건은 판매 페이지에서 확인하세요.")
    lines.append("")

    lines.append("## 이런 분께 추천")
    if bullets:
        for b in bullets[:5]:
            b = to_str(b)
            if b:
                lines.append(f"- {b}")
    else:
        lines.append("- 합리적인 가격대의 상품을 빠르게 비교하고 싶은 분")
        lines.append("- 실제 구매 페이지에서 최신 정보를 확인하고 싶은 분")
    lines.append("")

    lines.append("## 상품 스펙 한눈에 보기")
    if rows:
        lines.append("| 구분 | 상세 설명 |")
        lines.append("| --- | --- |")
        for r in rows[:8]:
            label = to_str(r.get("label")).replace("|", "\\|")
            detail = to_str(r.get("detail")).replace("|", "\\|")
            if label and detail:
                lines.append(f"| {label} | {detail} |")
        lines.append("")
    else:
        lines.append("상세 스펙은 판매 페이지에서 확인해 주세요.\n")

    lines.append("## 구매 전 체크포인트")
    lines.append("- 가격/재고/구성은 판매 페이지 기준으로 수시 변동될 수 있어요.")
    lines.append("- 배송/교환/반품 정책을 결제 전 꼭 확인하세요.\n")

    lines.append("## 자주 묻는 질문(FAQ)")
    if faq:
        for qa in faq[:3]:
            q = to_str(qa.get("q"))
            a = to_str(qa.get("a"))
            if q and a:
                lines.append(f"**Q. {q}**\n\nA. {a}\n")
    else:
        lines.append("**Q. 최신 가격은 어디서 확인하나요?**\n\nA. 아래 링크의 판매 페이지에서 최신 가격/혜택을 확인할 수 있어요.\n")

    lines.append(product_ld + "\n")
    lines.append(f"[← 상품 모아보기로 돌아가기]({BASE_PATH}/)\n")

    if product_url:
        lines.append(
            '<div style="margin-top:18px;padding:18px;border:2px solid #10b981;'
            'border-radius:14px;background:#ecfdf5;text-align:center;'
            'font-weight:700;color:#047857;font-size:15px;line-height:1.6;">'
        )
        lines.append("최신 혜택/재고는 판매 페이지에서 확인하세요.<br/>")
        lines.append(
            f'<a href="{product_url}" rel="nofollow sponsored" '
            'style="color:#065f46;text-decoration:none;font-weight:800;">'
            "최대 할인 확인하기 →"
            "</a>"
        )
        lines.append("</div>\n")

    return "\n".join(lines)


def build_index_markdown(items: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    lines.append("---")
    lines.append('title: "HAJINFAM 할인상품 모아보기"')
    lines.append('description: "인스타그램에서 소개된 상품을 번호로 빠르게 찾아볼 수 있는 모아보기 페이지입니다."')
    lines.append("sitemap: true")
    lines.append("noindex: false")
    lines.append("---\n")
    lines.append("# HAJINFAM 할인상품 모아보기\n")
    lines.append("인스타그램에서 보신 **번호 그대로** 검색하거나, 아래 목록에서 상품을 선택해 보세요.\n")

    lines.append("<style>")
    lines.append(r"""
.product-grid{
  display:grid;
  grid-template-columns:repeat(auto-fit,minmax(230px,1fr));
  gap:18px;
  margin-top:18px
}
.product-card{
  display:flex;
  flex-direction:column;
  border-radius:16px;
  background:#fff;
  border:1px solid #e5e7eb;
  box-shadow:0 10px 25px rgba(15,23,42,.06);
  text-decoration:none;
  color:#111827;
  overflow:hidden;
  transition:transform .12s ease,box-shadow .12s ease,border-color .12s ease
}
.product-card:hover{
  transform:translateY(-2px);
  border-color:#10b981;
  box-shadow:0 16px 35px rgba(16,185,129,.22)
}
.product-thumb{
  position:relative;
  width:100%;
  aspect-ratio:16/9;
  background:#f3f4f6;
}
.product-thumb img{
  width:100%;
  height:100%;
  object-fit:cover;
  display:block
}
.product-badge{
  position:absolute;
  top:10px;
  left:10px;
  display:inline-flex;
  align-items:center;
  justify-content:center;
  font-size:12px;
  font-weight:800;
  line-height:1;
  color:#fff;
  padding:6px 10px;
  border-radius:999px;
  background:rgba(17,24,39,.55);
  border:1px solid rgba(255,255,255,.25);
  backdrop-filter:blur(4px)
}
.product-body{
  display:flex;
  flex-direction:column;
  gap:6px;
  padding:12px 14px 14px;
  min-height:120px;
}
.product-title{
  font-size:15px;
  font-weight:800;
  line-height:1.35;
  margin:0
}
.product-summary{
  font-size:13px;
  color:#4b5563;
  line-height:1.5;
  margin:0;
  flex:1;
  display:-webkit-box;
  -webkit-line-clamp:2;
  -webkit-box-orient:vertical;
  overflow:hidden
}
.product-cta{
  margin-top:6px;
  font-size:13px;
  font-weight:700;
  color:#059669
}
.product-cta span{
  border-radius:999px;
  padding:6px 10px;
  background:#ecfdf5;
  display:inline-block
}
""")
    lines.append("</style>\n")

    lines.append('<div class="product-grid">')

    for it in sorted(items, key=lambda x: x["no"]):
        no = it["no"]
        title = html_escape(it.get("title", ""))
        summary = html_escape(it.get("summary", ""))
        permalink = it.get("permalink", "#")
        image_url = to_str(it.get("image_url") or it.get("imageUrl") or it.get("mainImage") or "")

        lines.append(f'<a class="product-card" href="{permalink}">')
        lines.append('  <div class="product-thumb">')
        if image_url:
            lines.append(f'    <img src="{image_url}" alt="{title}" loading="lazy" />')
        lines.append(f'    <div class="product-badge">{no}번</div>')
        lines.append('  </div>')

        lines.append('  <div class="product-body">')
        lines.append(f'    <h3 class="product-title">{title}</h3>')
        lines.append(f'    <p class="product-summary">{summary}</p>')
        lines.append('    <div class="product-cta"><span>자세히 보기 →</span></div>')
        lines.append('  </div>')
        lines.append('</a>')

    lines.append("</div>\n")
    return "\n".join(lines)


def gs_connect():
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(GSERVICE_KEY_PATH, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(GOOGLE_SHEET_ID)
    ws = sh.worksheet(GOOGLE_WORKSHEET_NAME)
    return ws


def sync_excel_seo_to_google_sheet(df: pd.DataFrame):
    """
    df: 엑셀에서 읽은 DataFrame (seoTitle/seoDescription/keywords/category 포함)
    - 'no' 기준으로 구글시트 행을 찾아 SEO 컬럼을 채움
    - 읽기: get_all_values() 1회
    - 쓰기: update_cells() 배치(청크)
    """
    ws = gs_connect()

    print("[GS] get_all_values start...", flush=True)
    values = ws.get_all_values()
    print(f"[GS] get_all_values done. rows={len(values)}", flush=True)

    if not values or len(values) < 2:
        raise RuntimeError("구글시트가 비어있거나 헤더(1행)가 없습니다.")

    headers = [str(h).strip() for h in values[0]]
    header_map = {h: idx for idx, h in enumerate(headers)}  # 0-based

    required = ["no", "seoTitle", "seoDescription", "keywords", "category"]
    for c in required:
        if c not in header_map:
            raise RuntimeError(f"구글시트 헤더(1행)에 '{c}' 컬럼이 없습니다. 현재 헤더: {headers}")

    no_col = header_map["no"]
    col_seoTitle = header_map["seoTitle"]
    col_seoDesc = header_map["seoDescription"]
    col_keys = header_map["keywords"]
    col_cat = header_map["category"]

    # no -> rownum(1-based)
    no_to_rownum: Dict[int, int] = {}
    for r_idx in range(1, len(values)):
        row_vals = values[r_idx]
        raw_no = str(row_vals[no_col]).strip() if no_col < len(row_vals) else ""
        try:
            n = int(raw_no)
        except Exception:
            continue
        no_to_rownum[n] = r_idx + 1

    cells_to_update: List[gspread.Cell] = []
    updated_rows = 0
    skipped = 0
    missing = 0

    df2 = df.copy()
    df2 = df2.fillna("")
    df2.columns = [str(c).strip() for c in df2.columns]

    for _, r in df2.iterrows():
        try:
            n = int(str(r.get("no", "")).strip())
        except Exception:
            continue

        if n not in no_to_rownum:
            missing += 1
            continue

        rownum = no_to_rownum[n]
        sheet_row_idx = rownum - 1
        row_vals = values[sheet_row_idx]

        def cur(col_index: int) -> str:
            return str(row_vals[col_index]).strip() if col_index < len(row_vals) else ""

        # 엑셀 값
        v_title = to_str(r.get("seoTitle"))
        v_desc = to_str(r.get("seoDescription"))
        v_keys = to_str(r.get("keywords"))
        v_cat = to_str(r.get("category"))

        cur_title = cur(col_seoTitle)
        cur_desc = cur(col_seoDesc)
        cur_keys = cur(col_keys)
        cur_cat = cur(col_cat)

        if GS_WRITE_MODE == "fill_empty":
            upd_title = v_title if (not cur_title and v_title) else None
            upd_desc = v_desc if (not cur_desc and v_desc) else None
            upd_keys = v_keys if (not cur_keys and v_keys) else None
            upd_cat = v_cat if (not cur_cat and v_cat) else None

            if not any([upd_title, upd_desc, upd_keys, upd_cat]):
                skipped += 1
                continue
        else:
            upd_title, upd_desc, upd_keys, upd_cat = v_title, v_desc, v_keys, v_cat

        if upd_title is not None:
            cells_to_update.append(gspread.Cell(rownum, col_seoTitle + 1, upd_title))
        if upd_desc is not None:
            cells_to_update.append(gspread.Cell(rownum, col_seoDesc + 1, upd_desc))
        if upd_keys is not None:
            cells_to_update.append(gspread.Cell(rownum, col_keys + 1, upd_keys))
        if upd_cat is not None:
            cells_to_update.append(gspread.Cell(rownum, col_cat + 1, upd_cat))

        updated_rows += 1

    CHUNK = 500
    for i in range(0, len(cells_to_update), CHUNK):
        ws.update_cells(cells_to_update[i:i + CHUNK], value_input_option="RAW")

    print(f"[GS DONE] rows_updated={updated_rows}, cells_written={len(cells_to_update)}, skipped={skipped}, missing_no={missing}", flush=True)


def main():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {EXCEL_PATH}")

    # 1) pandas로 읽기
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    if isinstance(df, dict):
        df = list(df.values())[0]
    df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")

    print("[INFO] rows:", len(df), "cols:", list(df.columns), flush=True)

    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            raise RuntimeError(f"엑셀에 필요한 컬럼이 없습니다: {col}")

    for c in SEO_COLUMNS:
        if c not in df.columns:
            df[c] = ""

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 2) openpyxl로 같은 시트 잡기
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]

    header_cells = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_map = {name: idx + 1 for idx, name in enumerate(header_cells)}

    # 없으면 헤더 추가
    for c in SEO_COLUMNS:
        if c not in col_map:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col).value = c
            col_map[c] = new_col

    generated_md = 0
    skipped_rows = 0
    updated_excel_rows = 0
    index_items: List[Dict[str, Any]] = []

    total = len(df)

    for i, r in df.iterrows():
        try:
            no = int(r["no"])
        except Exception:
            skipped_rows += 1
            continue

        name = to_str(r.get("productName"))
        product_url = to_str(r.get("productUrl"))

        # 진행 로그(매 row)
        print(f"[GEN] {i+1}/{total} no={no} name={shorten(name, 28)}", flush=True)

        filename = f"product_{no:03d}.md"
        filepath = os.path.join(OUTPUT_DIR, filename)
        permalink = f"{BASE_PATH}/product_{no:03d}/"

        # 재개 실행: 이미 있으면 스킵
        if SKIP_EXISTING_MD and os.path.exists(filepath):
            continue

        if not name or not product_url:
            skipped_rows += 1
            continue

        try:
            image_url = to_str(r.get("mainImage"))
            short_title = to_str(r.get("shortTitle"))
            base_desc = to_str(r.get("productDescription"))

            seo_title = to_str(r.get("seoTitle"))
            seo_desc = to_str(r.get("seoDescription"))
            keywords = to_str(r.get("keywords"))
            category = to_str(r.get("category"))

            needs_fill = (not seo_title) or (not seo_desc) or (not keywords) or (not category)

            # 비어있을 때만 GPT 호출
            if needs_fill:
                pack = generate_seo_pack(name, base_desc, keywords=keywords, category=category)
            else:
                pack = {
                    "seoTitle": seo_title,
                    "seoDescription": seo_desc,
                    "keywords": keywords,
                    "category": category,
                    "intro": "",
                    "bullets": [],
                    "rows": [],
                    "faq": [],
                }

            # 엑셀 빈칸만 채우기
            excel_row = i + 2
            updates = {}

            if not seo_title:
                seo_title = to_str(pack.get("seoTitle"))
                updates["seoTitle"] = seo_title
            if not seo_desc:
                seo_desc = to_str(pack.get("seoDescription"))
                updates["seoDescription"] = seo_desc
            if not keywords:
                keywords = to_str(pack.get("keywords"))
                updates["keywords"] = keywords
            if not category:
                category = to_str(pack.get("category"))
                updates["category"] = category

            if updates:
                for k, v in updates.items():
                    ws.cell(row=excel_row, column=col_map[k]).value = v
                updated_excel_rows += 1

            md = build_markdown(
                {
                    "no": no,
                    "productName": name,
                    "productUrl": product_url,
                    "mainImage": image_url,
                    "shortTitle": short_title,
                    "productDescription": base_desc,
                    "seoTitle": seo_title,
                    "seoDescription": seo_desc,
                    "keywords": keywords,
                    "category": category,
                },
                pack,
            )

            with open(filepath, "w", encoding="utf-8") as f:
                f.write(md)

            generated_md += 1
            index_items.append({
                "no": no,
                "title": short_title or name,
                "summary": shorten(seo_desc or base_desc, 46),
                "permalink": permalink,
                "image_url": image_url,
            })

            if STOP_AFTER is not None and generated_md >= STOP_AFTER:
                print(f"[STOP] generated_md reached {STOP_AFTER}.", flush=True)
                break

        except KeyboardInterrupt:
            raise
        except Exception:
            print(f"[ERR] no={no} (row={i+2})", flush=True)
            traceback.print_exc()
            skipped_rows += 1
            continue

    # index 생성 (이미 생성된 md도 포함하려면 별도 스캔이 필요하지만,
    # 현재는 이번 실행에서 생성/처리한 것 기준으로 index를 갱신)
    index_path = os.path.join(OUTPUT_DIR, "index.md")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(build_index_markdown(index_items))

    wb.save(EXCEL_PATH)
    print("[STEP] Excel saved.", flush=True)

    # 구글시트 동기화는 저장 후 재로드한 df로 (정확도)
    try:
        df_sync = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
        if isinstance(df_sync, dict):
            df_sync = list(df_sync.values())[0]
        df_sync.columns = [str(c).strip() for c in df_sync.columns]
        df_sync = df_sync.fillna("")

        print("[STEP] Google Sheet SEO sync start...", flush=True)
        sync_excel_seo_to_google_sheet(df_sync)
        print("[STEP] Google Sheet SEO sync done.", flush=True)
    except Exception as e:
        print(f"[WARN] Google Sheet sync failed: {e}", flush=True)

    print(f"[DONE] md 생성: {generated_md}개", flush=True)
    print(f"[DONE] 엑셀 SEO 자동 채움: {updated_excel_rows}행", flush=True)
    print(f"[DONE] 스킵된 행: {skipped_rows}행", flush=True)
    print(f"[DONE] index.md: {index_path}", flush=True)


if __name__ == "__main__":
    main()
